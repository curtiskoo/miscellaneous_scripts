#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun May 13 21:22:54 2018

@author: curtiskoo
"""

import openpyxl
from openpyxl import load_workbook
wb = load_workbook('testcolor.xlsx',data_only=True)
sh = wb['Sheet1']
cell = 'B3'
rgb = sh[cell].fill.start_color.rgb
#print(rgb)

for x in range(1,5):
    for l in "ABCDEFG":
        cell = "{}{}".format(l,x)
        rgb = sh[cell].fill.start_color.index
        print(rgb,l,x)
        if rgb == 2:
            print(l,x)

            
def white_cell(c,l=None,x=None):
    if l == None or x == None:
        cell = c
    else:
        cell = "{}{}".format(l,x)
    rgb = sh[cell].fill.start_color.rgb
    if rgb == '00000000':
        return True
    else:
        return False
    
white_cell('B1','C',1)


ws = wb.worksheets[0]
ws['A1'] = 'hellosir123'
wb.save('testcolor.xlsx')

ws['A1'].value
rows = ws.max_row
columns = ws.max_column

print(rows,columns)

def get_alpha(n):
    key = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    lst = []
    if n == 0:
        return
    tn = n
    while n > 26:
        tn = n // 26
        tr = n % 26
        if tr == 0 and tn == 27:
            #print(tn,tr)
            tn = 26
            tr = 26
        elif tr == 0 and tn != 27:
            #print(tn,tr)
            tn = tn - 1
            tr = 26
        lst = [tr] + lst
        n = tn
    lst = [tn] + lst
    
    s = ""
    for x in lst:
        s = s + key[x-1]
    return s, lst
get_alpha(53)